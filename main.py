from bs4 import BeautifulSoup
import requests
import json
import csv
from openpyxl import Workbook

#url = ['https://pt.wikipedia.org/wiki/Hidrog%C3%A9nio', 'https://pt.wikipedia.org/wiki/L%C3%ADtio']

def fetchLinks():
    linkURL = 'https://pt.wikipedia.org/wiki/Tabela_peri%C3%B3dica'

    page = requests.get(linkURL)

    soup = BeautifulSoup(page.text, 'html.parser')

    tables = soup.find_all('table') 

    a = tables[0].find_all('a')

    #print(type(a))
    #print(len(a))

    output = []

    for i, rows in enumerate(a):
        if('/wiki/' in rows['href'] and 'Período' not in rows['title'] and 'Grupo' not in rows['title'] and i > 18):
            output.append('https://pt.wikipedia.org' + rows['href'])
            #print(f"{i} | {rows}")

    return output

def fetchStuff(target_URL):    
    page = requests.get(target_URL)

    soup = BeautifulSoup(page.text, 'html.parser')

    table = soup.find('table', class_='infobox_v2')

    tr = table.find_all('tr')

    data_array = []

    #target_IDS = [21, 22, 23, 24, 25, 28, 29, 30, 31, 32, 33, 34, 36, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 53, 54, 55, 56, 57, 58, 61, 63, 64, 65, 66]
    target_IDS = [21, 22, 23, 24, 25, 28, 29, 30, 31, 32, 33, 34, 36, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48]

    for i, row in enumerate(tr):
        row_data = row.find_all('td')
        processsed_row_data = [data.text.strip() for data in row_data]
        if(True):
            #print(f"ROW {i}: {processsed_row_data}")
            data_array.append(processsed_row_data)
    
    processed_data_array = []

    for i, id in enumerate(target_IDS):
        processed_data_array.append(data_array[id])
        #print(f"{i} | {data_array[id]}")

    #print(processed_data_array[10][1])
    electrons = []

    for electron in processed_data_array[10][1].split(','):
        processed_electron = electron.replace('(ver imagem)', '').strip()
        if(processed_electron.isdigit()):
            electrons.append(int(processed_electron))

    #print(electrons)
    
    obj = {
        'nome': processed_data_array[0][1].split(',')[0],
        'simbolo': processed_data_array[0][1].split(',')[1].strip(),
        'numero': int(processed_data_array[0][1].split(',')[2].strip()),
        'serie_quimica': processed_data_array[1][1],
        'grupo': processed_data_array[2][1].split(',')[0],
        'periodo': int(processed_data_array[2][1].split(',')[1]),
        'bloco': processed_data_array[2][1].split(',')[2].strip(),
        'densidade_dureza': ' '.join(processed_data_array[3][1].split()),
        #'dureza': dureza,
        'numero_CAS': processed_data_array[4][1],
        'massa_atomica': processed_data_array[5][1],
        'raio_atomico_calculado': processed_data_array[6][1],
        'raio_covalente': processed_data_array[7][1],
        'raio_de_van_der_waals': processed_data_array[8][1],
        'configuracao_eletronica': processed_data_array[9][1],
        'eletrons': electrons,
        'estado_de_oxidacao': processed_data_array[11][1],
        'estrutura_cristalina': processed_data_array[12][1],
        'estado_da_materia': processed_data_array[13][1],
        'ponto_de_fusao': processed_data_array[14][1],
        'ponto_de_ebulicao': processed_data_array[15][1],
        'entalpia_de_fusao': processed_data_array[16][1],
        'entalpia_de_vaporizacao':processed_data_array[17][1],
        #'temperatura_critica': processed_data_array[18][1],
        #'pressao_critica': processed_data_array[19][1],
        'volume_molar': processed_data_array[20][1],
        'pressao_de_vapor': processed_data_array[21][1],
        'velocidade_do_som': processed_data_array[22][1],
        'classe_magnetica': processed_data_array[23][1]
    }

    #print(obj)

    return obj

#fetchStuff(url[0])

urls = fetchLinks()

print(f"URLs amount: {len(urls)}")

objs = []

limit = 82

for i, url in enumerate(urls):
    print(f"Pulling URL #{i} | {url}")
    current_obj = fetchStuff(url)
    if(current_obj['numero'] <= limit):
        objs.append(current_obj)

    #print(f"Numero Atomico: {current_obj['numero']}")
    #break

objs.sort(key=lambda x: x['numero'], reverse=False)

#print(objs)
def saveJSON():
    json_dump = json.dumps(objs, indent=3, ensure_ascii=False)

    #print(json_dump)

    with open('data.json', 'w', encoding='utf8') as outfile:
        outfile.write(json_dump)
    
    print('JSON saved')

headers = [
    'nome',
    'simbolo',
    'numero',
    'serie_quimica',
    'grupo',
    'periodo',
    'bloco',
    'densidade_dureza',
    'numero_CAS',
    'massa_atomica',
    'raio_atomico_calculado',
    'raio_covalente',
    'raio_de_van_der_waals',
    'configuracao_eletronica',
    'eletrons[0]',
    'eletrons[1]',
    'eletrons[2]',
    'eletrons[3]',
    'eletrons[4]',
    'eletrons[5]',
    'estado_de_oxidacao',
    'estrutura_cristalina',
    'estado_da_materia',
    'ponto_de_fusao',
    'ponto_de_ebulicao',
    'entalpia_de_fusao',
    'entalpia_de_vaporizacao',
    'volume_molar',
    'pressao_de_vapor',
    'velocidade_do_som',
    'classe_magnetica'
]

def saveCSV():
    with open('data.csv', mode='w', encoding='utf8') as outfile:

        filewriter = csv.writer(outfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

        filewriter.writerow(headers)

        for obj in objs:
            electrons = []
            max_num = len(obj['eletrons'])
            for i in range(6):
                if(i < max_num):
                    electrons.append(obj['eletrons'][i])
                else:
                    electrons.append(0)

            filewriter.writerow([
                obj['nome'],
                obj['simbolo'],
                obj['numero'],
                obj['serie_quimica'],
                obj['grupo'],
                obj['periodo'],
                obj['bloco'],
                obj['densidade_dureza'],
                obj['numero_CAS'],
                obj['massa_atomica'],
                obj['raio_atomico_calculado'],
                obj['raio_covalente'],
                obj['raio_de_van_der_waals'],
                obj['configuracao_eletronica'],
                electrons[0],
                electrons[1],
                electrons[2],
                electrons[3],
                electrons[4],
                electrons[5],
                obj['estado_de_oxidacao'],
                obj['estrutura_cristalina'],
                obj['estado_da_materia'],
                obj['ponto_de_fusao'],
                obj['ponto_de_ebulicao'],
                obj['entalpia_de_fusao'],
                obj['entalpia_de_vaporizacao'],
                obj['volume_molar'],
                obj['pressao_de_vapor'],
                obj['velocidade_do_som'],
                obj['classe_magnetica'],
            ])
    print('CSV saved')

def saveEXCEL():

    filename = 'data.xlsx'

    workbook = Workbook()
    sheet = workbook.active

    #SET HEADERS
    for i, header in enumerate(headers):
        sheet.cell(row=1, column=(i+1)).value = header

    #ADD DATA
    for i, obj in enumerate(objs):
        electrons = []
        max_num = len(obj['eletrons'])

        '''
        for item in obj.items():
            print(item[1])
        '''

        for e in range(6):
            if(e < max_num):
                electrons.append(obj['eletrons'][e])
            else:
                electrons.append(0)
        
        sheet.cell(row=(i+2), column=1).value = obj['nome']
        sheet.cell(row=(i+2), column=2).value = obj['simbolo']
        sheet.cell(row=(i+2), column=3).value = obj['numero']
        sheet.cell(row=(i+2), column=4).value = obj['serie_quimica']
        sheet.cell(row=(i+2), column=5).value = obj['grupo']
        sheet.cell(row=(i+2), column=6).value = obj['periodo']
        sheet.cell(row=(i+2), column=7).value = obj['bloco']
        sheet.cell(row=(i+2), column=8).value = obj['densidade_dureza']
        sheet.cell(row=(i+2), column=9).value = obj['numero_CAS']
        sheet.cell(row=(i+2), column=10).value = obj['massa_atomica']
        sheet.cell(row=(i+2), column=10).value = obj['raio_atomico_calculado']
        sheet.cell(row=(i+2), column=11).value = obj['raio_covalente']
        sheet.cell(row=(i+2), column=12).value = obj['raio_de_van_der_waals']
        sheet.cell(row=(i+2), column=13).value = obj['nome']
        sheet.cell(row=(i+2), column=14).value = obj['configuracao_eletronica']
        sheet.cell(row=(i+2), column=15).value = electrons[0]
        sheet.cell(row=(i+2), column=16).value = electrons[1]
        sheet.cell(row=(i+2), column=17).value = electrons[2]
        sheet.cell(row=(i+2), column=18).value = electrons[3]
        sheet.cell(row=(i+2), column=19).value = electrons[4]
        sheet.cell(row=(i+2), column=20).value = electrons[5]
        sheet.cell(row=(i+2), column=21).value = obj['estado_de_oxidacao']
        sheet.cell(row=(i+2), column=22).value = obj['estrutura_cristalina']
        sheet.cell(row=(i+2), column=23).value = obj['estado_da_materia']
        sheet.cell(row=(i+2), column=24).value = obj['ponto_de_fusao']
        sheet.cell(row=(i+2), column=25).value = obj['ponto_de_ebulicao']
        sheet.cell(row=(i+2), column=26).value = obj['entalpia_de_fusao']
        sheet.cell(row=(i+2), column=27).value = obj['entalpia_de_vaporizacao']
        sheet.cell(row=(i+2), column=28).value = obj['volume_molar']
        sheet.cell(row=(i+2), column=29).value = obj['pressao_de_vapor']
        sheet.cell(row=(i+2), column=30).value = obj['velocidade_do_som']
        sheet.cell(row=(i+2), column=31).value = obj['classe_magnetica']

    sheet.freeze_panes = "A2"        

    workbook.save(filename=filename)

    print('EXCEL saved')

saveJSON()
saveCSV()
saveEXCEL()